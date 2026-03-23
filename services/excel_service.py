import io
import pandas as pd
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict

logger = logging.getLogger(__name__)


def _apply_trend_colors(worksheet, trend_column_letter: str, max_row: int):
    """Применяет цветовую заливку для ячеек с трендом"""
    # Цвета
    up_fill = PatternFill("solid", fgColor="FFC6EFCE")      # зеленый для восходящего
    down_fill = PatternFill("solid", fgColor="FFFFC7CE")    # красный для нисходящего
    stable_fill = PatternFill("solid", fgColor="FFFFEB9C")  # желтый для стабильного
    na_fill = PatternFill("solid", fgColor="FFE0E0E0")      # серый для недостаточно данных
    
    for row in range(4, max_row + 1):
        cell = worksheet.cell(row=row, column=trend_column_letter)
        value = str(cell.value).lower() if cell.value else ""
        
        if value in ["вверх", "восходящий"]:
            cell.fill = up_fill
            cell.value = "↑ вверх"
        elif value in ["вниз", "нисходящий"]:
            cell.fill = down_fill
            cell.value = "↓ вниз"
        elif value in ["ровно", "стабильный"]:
            cell.fill = stable_fill
            cell.value = "→ ровно"
        else:
            cell.fill = na_fill
            cell.value = "? нет данных"


def _apply_fixed_column_widths_like_example(worksheet):
    """Фиксированные ширины столбцов"""
    widths = {
        "A": 15.86,   # Ссылка на Ozon
        "B": 19.0,    # Категория
        "C": 19.14,   # Название товара
        "D": 11.0,    # Цена, р
        "E": 19.0,    # Выручка за 30 дней
        "F": 14.0,    # Кол-во конкурентов
        "G": 13.0,    # Тренд
        "H": 15.0,    # Кол-во к закупке
        "I": 10.0,    # Себестоимость
        "J": 12.0,    # % Комиссии
        "K": 12.14,   # Комиссия, р
        "L": 14.0,    # Логистика
        "M": 13.0,    # Эквайринг, р
        "N": 13.0,    # Всего расходы на ед, р
        "O": 13.0,    # Закуп итого, р
        "P": 10.57,   # Прибыль до налогов, р
        "Q": 8.0,     # Прибыль на партию, р
        "R": 8.0,     # План по выручке, р
        "S": 8.0,     # Налоги, р
        "T": 8.0,     # Прибыль после налогов, р
        "U": 8.0,     # Маржа, %
        "V": 8.0,     # ROI, %
    }
    
    for col_letter, width in widths.items():
        try:
            worksheet.column_dimensions[col_letter].width = width
        except Exception:
            pass


def _apply_header_style(worksheet, max_row: int):
    """Применяет стиль для заголовков"""
    fill = PatternFill("solid", fgColor="FFEFEFEF")
    font = Font(bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=3, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _apply_user_input_column_style(worksheet, header_names: list[str], max_row: int):
    """Excel-like light green fill for input columns"""
    green_fill = PatternFill("solid", fgColor="FFC6EFCE")
    alignment = Alignment(vertical="center")

    headers = [worksheet.cell(3, c).value for c in range(1, worksheet.max_column + 1)]
    for name in header_names:
        if name not in headers:
            continue
        col = headers.index(name) + 1
        hcell = worksheet.cell(3, col)
        hcell.fill = green_fill
        hcell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in range(4, max_row + 1):
            cell = worksheet.cell(row, col)
            cell.fill = green_fill
            cell.alignment = alignment


def _add_tax_rate_cell(worksheet):
    """Добавляет ячейку с налоговой ставкой в H1 и H2"""
    # Ячейка с названием
    tax_label = worksheet.cell(row=1, column=8)  # H1
    tax_label.value = "Налоговая ставка"
    tax_label.fill = PatternFill("solid", fgColor="FFC6EFCE")
    tax_label.font = Font(bold=True)
    tax_label.alignment = Alignment(horizontal="center")
    
    # Ячейка со значением ставки
    tax_rate_cell = worksheet.cell(row=2, column=8)  # H2
    tax_rate_cell.value = 0.06
    tax_rate_cell.number_format = '0%'
    tax_rate_cell.fill = PatternFill("solid", fgColor="FFC6EFCE")
    tax_rate_cell.alignment = Alignment(horizontal="center")
    
    return get_column_letter(8)  # H


def create_excel_report(results: List[Dict]) -> io.BytesIO:
    """Создает Excel отчет с результатами анализа"""
    if not results:
        df = pd.DataFrame([{'Статус': 'Нет данных'}])
    else:
        if results:
            sample = results[0]
            logger.info(f"Пример данных: категория={sample.get('category')}, комиссия={sample.get('commission')}")
        
        df = pd.DataFrame(results)

        # Приводим входные ключи к новой структуре отчёта
        df = df.rename(columns={
            "category": "Категория",
            "name": "Название товара",
            "price": "Цена, р",
            "revenue": "Выручка за 30 дней",
            "competitors": "Кол-во конкурентов",
            "trend": "Тренд",
        })

        # Удаляем столбцы Бренд и Продавец
        for drop_col in ("brand", "seller"):
            if drop_col in df.columns:
                df = df.drop(columns=[drop_col])

        if "url" in df.columns:
            df["Ссылка на Ozon"] = df["url"]
            df = df.drop(columns=["url"])

        # Пользовательские/расчётные колонки
        df["Кол-во к закупке"] = ""
        df["Себестоимость"] = ""
        df["% Комиссии"] = df.get("commission_percent", "")
        df["Комиссия, р"] = df.get("commission", 0)
        df["Логистика"] = df.get("logistics", 0)
        df["Эквайринг, р"] = ""
        df["Всего расходы на ед, р"] = ""
        df["Закуп итого, р"] = ""
        df["Прибыль до налогов, р"] = ""
        df["Прибыль на партию, р"] = ""
        df["План по выручке, р"] = ""
        df["Налоги, р"] = ""
        df["Прибыль после налогов, р"] = ""
        df["Маржа, %"] = ""
        df["ROI, %"] = ""

        # Удаляем временные столбцы
        for col in ["commission", "commission_percent", "logistics"]:
            if col in df.columns:
                df = df.drop(columns=[col])

        # Итоговый порядок столбцов
        col_order = [
            "Ссылка на Ozon",
            "Категория",
            "Название товара",
            "Цена, р",
            "Выручка за 30 дней",
            "Кол-во конкурентов",
            "Тренд",
            "Кол-во к закупке",
            "Себестоимость",
            "% Комиссии",
            "Комиссия, р",
            "Логистика",
            "Эквайринг, р",
            "Всего расходы на ед, р",
            "Закуп итого, р",
            "Прибыль до налогов, р",
            "Прибыль на партию, р",
            "План по выручке, р",
            "Налоги, р",
            "Прибыль после налогов, р",
            "Маржа, %",
            "ROI, %",
        ]
        df = df[[c for c in col_order if c in df.columns]]

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Результаты анализа')

        worksheet = writer.sheets['Результаты анализа']
        
        # Добавляем ячейку с налоговой ставкой
        tax_col_letter = _add_tax_rate_cell(worksheet)
        
        # Вставляем 2 пустые строки сверху для налоговой ставки
        worksheet.insert_rows(0, amount=2)
        
        # Заголовки теперь на строке 3
        for col_idx, cell in enumerate(df.columns, 1):
            worksheet.cell(row=3, column=col_idx, value=cell)
        
        # Копируем данные в строки начиная с 4
        for row_idx, row_data in enumerate(df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Получаем индексы колонок
        headers = [worksheet.cell(3, c).value for c in range(1, worksheet.max_column + 1)]
        
        def col_idx(name: str):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        # Получаем индексы всех колонок
        c_link = col_idx("Ссылка на Ozon")
        c_price = col_idx("Цена, р")
        c_qty = col_idx("Кол-во к закупке")
        c_cogs = col_idx("Себестоимость")
        c_commission_percent = col_idx("% Комиссии")
        c_commission_rub = col_idx("Комиссия, р")
        c_log = col_idx("Логистика")
        c_acq = col_idx("Эквайринг, р")
        c_total = col_idx("Всего расходы на ед, р")
        c_buy_total = col_idx("Закуп итого, р")
        c_profit_before_tax = col_idx("Прибыль до налогов, р")
        c_profit_batch = col_idx("Прибыль на партию, р")
        c_plan_revenue = col_idx("План по выручке, р")
        c_tax = col_idx("Налоги, р")
        c_profit_after_tax = col_idx("Прибыль после налогов, р")
        c_margin = col_idx("Маржа, %")
        c_roi = col_idx("ROI, %")
        c_rev30 = col_idx("Выручка за 30 дней")
        c_trend = col_idx("Тренд")

        max_row = worksheet.max_row
        tax_rate_ref = f"{tax_col_letter}2"

        # Заполняем формулы для каждой строки
        for row in range(4, max_row + 1):
            # Кликабельная ссылка
            if c_link is not None:
                cell = worksheet.cell(row=row, column=c_link)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    url = cell.value.replace('"', '""')
                    cell.value = f'=HYPERLINK("{url}","{url}")'
                    cell.style = "Hyperlink"

            # Эквайринг = Цена * 1.5%
            if c_acq is not None and c_price is not None:
                worksheet.cell(
                    row=row,
                    column=c_acq,
                    value=f"={get_column_letter(c_price)}{row}*0.015",
                )

            # Всего расходы на ед = Себестоимость + Комиссия + Логистика + Эквайринг
            expense_columns = []
            if c_cogs is not None:
                expense_columns.append(get_column_letter(c_cogs))
            if c_commission_rub is not None:
                expense_columns.append(get_column_letter(c_commission_rub))
            if c_log is not None:
                expense_columns.append(get_column_letter(c_log))
            if c_acq is not None:
                expense_columns.append(get_column_letter(c_acq))
            
            if c_total is not None and expense_columns:
                formula = "+".join([f"{col}{row}" for col in expense_columns])
                worksheet.cell(row=row, column=c_total, value=f"={formula}")

            # Закуп итого = Кол-во к закупке * Себестоимость
            if c_buy_total is not None and c_qty is not None and c_cogs is not None:
                worksheet.cell(
                    row=row,
                    column=c_buy_total,
                    value=f"={get_column_letter(c_qty)}{row}*{get_column_letter(c_cogs)}{row}",
                )

            # Прибыль до налогов = Цена - Всего расходы на ед
            if c_profit_before_tax is not None and c_price is not None and c_total is not None:
                worksheet.cell(
                    row=row,
                    column=c_profit_before_tax,
                    value=f"={get_column_letter(c_price)}{row}-{get_column_letter(c_total)}{row}",
                )

            # Прибыль на партию = Прибыль до налогов * Кол-во к закупке
            if c_profit_batch is not None and c_profit_before_tax is not None and c_qty is not None:
                worksheet.cell(
                    row=row,
                    column=c_profit_batch,
                    value=f"={get_column_letter(c_profit_before_tax)}{row}*{get_column_letter(c_qty)}{row}",
                )

            # Налоги = Прибыль до налогов * Ставка налога
            if c_tax is not None and c_profit_before_tax is not None:
                worksheet.cell(
                    row=row,
                    column=c_tax,
                    value=f"={get_column_letter(c_profit_before_tax)}{row}*{tax_rate_ref}",
                )

            # Прибыль после налогов = Прибыль до налогов - Налоги
            if c_profit_after_tax is not None and c_profit_before_tax is not None and c_tax is not None:
                worksheet.cell(
                    row=row,
                    column=c_profit_after_tax,
                    value=f"={get_column_letter(c_profit_before_tax)}{row}-{get_column_letter(c_tax)}{row}",
                )

            # Маржа (%) = Прибыль после налогов / Цена * 100
            if c_margin is not None and c_profit_after_tax is not None and c_price is not None:
                worksheet.cell(
                    row=row,
                    column=c_margin,
                    value=f"=IF({get_column_letter(c_price)}{row}>0,"
                          f"{get_column_letter(c_profit_after_tax)}{row}/{get_column_letter(c_price)}{row}*100,\"\")",
                )

            # ROI (%) = Прибыль после налогов / Всего расходы на ед * 100
            if c_roi is not None and c_profit_after_tax is not None and c_total is not None:
                worksheet.cell(
                    row=row,
                    column=c_roi,
                    value=f"=IF({get_column_letter(c_total)}{row}>0,"
                          f"{get_column_letter(c_profit_after_tax)}{row}/{get_column_letter(c_total)}{row}*100,\"\")",
                )

        # Применяем цвета для тренда
        if c_trend is not None:
            trend_col_letter = get_column_letter(c_trend)
            _apply_trend_colors(worksheet, trend_col_letter, max_row)

        # Форматы чисел
        rub_fmt = '#,##0\\ _₽'
        pct_fmt = '0%'
        
        for row in range(4, max_row + 1):
            if c_price is not None:
                worksheet.cell(row=row, column=c_price).number_format = rub_fmt
            if c_rev30 is not None:
                worksheet.cell(row=row, column=c_rev30).number_format = rub_fmt
            if c_cogs is not None:
                worksheet.cell(row=row, column=c_cogs).number_format = rub_fmt
            if c_commission_rub is not None:
                worksheet.cell(row=row, column=c_commission_rub).number_format = rub_fmt
            if c_log is not None:
                worksheet.cell(row=row, column=c_log).number_format = rub_fmt
            if c_acq is not None:
                worksheet.cell(row=row, column=c_acq).number_format = rub_fmt
            if c_total is not None:
                worksheet.cell(row=row, column=c_total).number_format = rub_fmt
            if c_buy_total is not None:
                worksheet.cell(row=row, column=c_buy_total).number_format = rub_fmt
            if c_profit_before_tax is not None:
                worksheet.cell(row=row, column=c_profit_before_tax).number_format = rub_fmt
            if c_profit_batch is not None:
                worksheet.cell(row=row, column=c_profit_batch).number_format = rub_fmt
            if c_plan_revenue is not None:
                worksheet.cell(row=row, column=c_plan_revenue).number_format = rub_fmt
            if c_tax is not None:
                worksheet.cell(row=row, column=c_tax).number_format = rub_fmt
            if c_profit_after_tax is not None:
                worksheet.cell(row=row, column=c_profit_after_tax).number_format = rub_fmt
            if c_margin is not None:
                worksheet.cell(row=row, column=c_margin).number_format = pct_fmt
            if c_roi is not None:
                worksheet.cell(row=row, column=c_roi).number_format = pct_fmt
            if c_commission_percent is not None:
                worksheet.cell(row=row, column=c_commission_percent).number_format = pct_fmt

        # Применяем стили
        _apply_header_style(worksheet, max_row)
        _apply_user_input_column_style(worksheet, ["Кол-во к закупке", "Себестоимость"], max_row)
        _apply_fixed_column_widths_like_example(worksheet)

    out.seek(0)
    return out


def create_category_template(categories):
    """Создает Excel шаблон со ВСЕМИ категориями"""
    if not categories:
        return None

    data = []
    for i, cat in enumerate(categories, 1):
        path = cat.get('path', '')
        name = cat.get('name', '')

        path_parts = path.split('/') if path else []

        main_category = path_parts[0] if len(path_parts) > 0 else name
        subcategory = '/'.join(path_parts[1:]) if len(path_parts) > 1 else ''

        data.append({
            '№': i,
            'Категория': name,
            'Основная категория': main_category,
            'Подкатегория': subcategory,
            'Полный путь': path,
            'Выбрать': 'НЕТ'
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Категории')

        worksheet = writer.sheets['Категории']
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 50
        worksheet.column_dimensions['E'].width = 80
        worksheet.column_dimensions['F'].width = 10

        worksheet['G1'] = '📌 ИНСТРУКЦИЯ:'
        worksheet['G2'] = '1. Всего категорий: ' + str(len(categories))
        worksheet['G3'] = '2. Поставьте "ДА" в колонке "Выбрать" для нужных категорий'
        worksheet['G4'] = '3. Можно фильтровать по основной категории'
        worksheet['G5'] = '4. Сохраните файл и загрузите обратно'

    output.seek(0)
    return output


def parse_categories_from_excel(file_bytes, apply_exclusions=False):
    """Парсит загруженный Excel файл и возвращает список выбранных категорий"""
    try:
        xls = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)

        def norm_col(c) -> str:
            return str(c).strip().lower()

        def pick_sheet():
            for _, df in xls.items():
                if df is None or df.empty:
                    continue
                cols_norm = {norm_col(c): c for c in df.columns}
                has_full_path = "полный путь" in cols_norm
                has_pair = ("категория" in cols_norm and "путь" in cols_norm)
                if has_full_path or has_pair:
                    return df, cols_norm
            return None, None

        df, cols_norm = pick_sheet()
        if df is None:
            return None

        selected = []

        choose_column = cols_norm.get("выбрать")

        if choose_column:
            for _, row in df.iterrows():
                choose_value = str(row.get(choose_column, "")).strip().lower()
                if choose_value in {"да", "yes", "1", "true", "y"}:
                    full_path_col = cols_norm.get("полный путь")
                    if full_path_col is not None:
                        path = row.get(full_path_col)
                        name = path.split('/')[-1] if path else ''
                    else:
                        name = row.get(cols_norm.get("категория"))
                        path = row.get(cols_norm.get("путь"))

                    if apply_exclusions:
                        from categories import is_allowed_category
                        if not is_allowed_category(name, path):
                            continue

                    selected.append({
                        'name': str(name or ""),
                        'path': str(path or ""),
                        'user_defined': True
                    })
        else:
            for _, row in df.iterrows():
                full_path_col = cols_norm.get("полный путь")
                if full_path_col is not None:
                    path = row.get(full_path_col)
                    name = path.split('/')[-1] if path else ''
                else:
                    name = row.get(cols_norm.get("категория"))
                    path = row.get(cols_norm.get("путь"))

                if apply_exclusions:
                    from categories import is_allowed_category
                    if not is_allowed_category(name, path):
                        continue

                selected.append({
                    'name': str(name or ""),
                    'path': str(path or ""),
                    'user_defined': True
                })

        return selected

    except Exception as e:
        print(f"Ошибка парсинга Excel: {e}")
        return None